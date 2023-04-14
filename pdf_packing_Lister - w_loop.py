import re
from pdfminer.high_level import extract_pages, extract_text
import geonamescache
import pycountry
import pypdf
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment

gc = geonamescache.GeonamesCache()
cities = gc.get_cities()
state_list = list(pycountry.countries) 
city_names = set(city['name'] for city in cities.values())

"""for page_layout in extract_pages("invoice.pdf"):
    for element in page_layout:
        print(element)"""



while True:
    try:
        ask_for_name = input("What pdf file do you want to load: ")
        inv_header = ask_for_name + ".pdf"        
        text = extract_text(inv_header)
        break
    except Exception as chyba:
        print(chyba)


delivery = text.find("Delivery")
def end_of_delivery():
    end1 = text.find("Payment")
    end2 = text.find("For\xa0custom")
    len1 = len(text[delivery:end1-1])
    len2 = len(text[delivery:end2-1])
    if len1 < len2 and len1 > 50:
        print("Returning end1 - Payment")
        return end1
    elif len2 < len1 and len2 > 50:
        print("Returning end2 - For custom")
        return end2
    else:
        return end1
end = end_of_delivery()    


delivery_adress = text[delivery:end-1]
delivery_adress = delivery_adress.replace("Delivery adress:","")
delivery_adress = delivery_adress.lstrip("\n")

incoterm_list = ["EXW","FCA","FOB","CFR","CIF","CIP""CPT","DAP","DAT","DDP"]

def incoterm_overlook(incoterm_list):
    for i in incoterm_list:
        found = int(text.find(i))
        if found != -1:
            if i == "DDP" and text[found+4] == "V":
                return "DDP VAT Unpaid"
            else:
                return i

def find_city(delivery_adress):
    tokenized = delivery_adress.split()
    if "Asahi" in tokenized:
        tokenized.remove("Asahi")
    for i in tokenized:
        if gc.get_cities_by_name(i):
            return i
    
    

def find_state_code():
    tokenized = delivery_adress.split()
    if "AS" in tokenized:
        tokenized.remove("AS")
    for i in tokenized:
        if i in [country.alpha_2 for country in pycountry.countries]:
            return i 
      
    
founded_inco = incoterm_overlook(incoterm_list)
founded_city = find_city(delivery_adress)
founded_state = find_state_code()


pdf_reader = pypdf.PdfReader(inv_header)



# Get the number of pages in the PDF file
num_pages = len(pdf_reader.pages)

# Loop over each page in the PDF file
def print_pages():
    for page_num in range(num_pages):
        # Get the page object for the current page
        page = pdf_reader.pages[page_num]

        # Extract the text content from the page
        page_text = page.extract_text()

        # Print the text content to the console
        var1 = page_text.find("GOODS")
        var2 = page_text.find("PRODUCTS")
        if var1 > var2:
            table_start = var1
        else:
            table_start = var2
            
        table_end = page_text.find("OTHERS")
        table = page_text[table_start:table_end]
        """if table_start == var1:
            table = [i for i in table if i != "PRODUCTS"]
        if table_start == var2:
            table = [i for i in table if i != "GOODS"]   """ 
        return table,page_text
    



# Define the regular expression patterns
zb_pattern = r"ZB\w+"
quant_pattern = r"pc\s*\d{1,3}(?:\s+\d{3})*" # matches zero or more groups of one or more whitespace characters followed by exactly three digits.
desc_pattern = r"%\s*\d{1,3}(?:\s+\d{3})*,\d{2}" # This pattern matches a percentage sign followed by a space, and then 1 to 5 digits, a comma, and 2 more digits. It can be used to match strings like "% 465,12".
netto_pattern = r"%\s*\d{1,3}(?:\s+\d{3})*,\d{2}" # + ,00
brutto_pattern = r"(?<=Packaging:).*?\n(\d[\d,. ]*)\s*Total Gross Weight"
inv_pattern = r"\d{2}03\w+"  
dl_pattern = r"\d{2}(?:SL|MH)01\w+"
order_pattern = r"\d{2}[A-Z0-9]{4}0\d{9}INVOICE"


table,page_text = print_pages()

print(table)
# Find all matches of the code pattern in the input string
codes = re.findall(zb_pattern, table)
quantity_list = re.findall(quant_pattern, table)
quantity_list = [quantity.replace('pc', '').replace('\xa0', '').strip() for quantity in quantity_list]
brutto_list = re.findall(brutto_pattern,page_text)
description = re.findall(desc_pattern, table)
#description = [desc.replace("\xa0"," ").strip() for desc in description] # this will remove \xa0 from the string but it will not find it afterwards
netto = re.findall(netto_pattern,table)
netto = [net.replace("%","").replace('\xa0', '') for net in netto]


def inv_lookup(what_to_look_for, text):
    findings = re.findall(what_to_look_for, text)
    return findings

invoice = re.search(inv_pattern, page_text)
invoice = invoice.group(0)

dl_note = inv_lookup(dl_pattern,page_text)
order_num = inv_lookup(order_pattern, page_text)
order_num = [order.replace("INVOICE","") for order in order_num]

def last_element(table, pattern):
    row_table = table.split('\n')
    end_list = []
    for row in row_table:
        if re.search(pattern, row):
            end = row.split()[-1]
            end_list.append(end)
    return end_list, row_table

end_list,row_table = last_element(table,zb_pattern)

print("\n\n\n")
def solidity_check(founded_city, founded_state):
    if founded_city is None or founded_state is None:
        if founded_city is None:
            print("City is missing")
            founded_city = input("Type the name of the city: ")
        if founded_state is None:
            print("Country code is missing")
            founded_state = input("Type the code of the country: ")
        return solidity_check(founded_city, founded_state)
    
    print("City and Country code are OK")
    return founded_city, founded_state
founded_city,founded_state = solidity_check(founded_city,founded_state)


# Print the results
print("\n\n\n")
print("Founded incoterm: ",founded_inco)
print("Founded city: ",founded_city)
print("Founded state: ",founded_state)
print("Codes: ", codes)
print("Quantity: ", quantity_list)
print("Start Product description: ", description)
print("Netto list is: ", netto)
print("Brutto list is: ", brutto_list)
#print("End list : ",end_list)  
print("Invoice number is: ", invoice)
print("Delivery note number is:", dl_note)
print("Order number is: ", order_num)

#perc_netto = ["%" + x for x in netto] # can replicate like ["%" + x,   for x in netto]
#print("percent netto: ",perc_netto)

def find_description():
    description_list = []
    for start, end in zip(description, end_list):
        start_pos = table.find(start) + len(start)
        end_pos = table.find("\n", start_pos) # finds the index of the first occurrence of end after the start substring.
        desc_row=table[start_pos:end_pos]
        description_list.append(desc_row)
    return description_list

description_list = find_description()
description_list = [desc.lstrip("\xa0").lstrip(" ") for desc in description_list]
#print(repr(description_list))
#print(description_list)
        
print("\n\n\n" )
#print(page_text)
#print(table)
#print(text)
#print(repr(delivery_adress))

#-----------------------------------------------------------------------------
#-----------------------------------------------------------------------------
#-----------------------------------------------------------------------------
#Quick Volume calculation PART
response = ["y","yes","yap","yeah",","]
# Create an empty DataFrame
quant = []
length = []
width = []
height = []
weight= []

named_columns ={"Quantity [0]": quant,
                "Length [1]": length,
                "Width [2]": width,
                "Height [3]": height}
named_columns_w_weight = {"Quantity [0]": quant,
                          "Length [1]": length,
                          "Width [2]": width,
                          "Height [3]": height,
                          "Weight [4]": weight,
                          "----": []}
df = pd.DataFrame(named_columns)
df2 = pd.DataFrame(named_columns_w_weight)



# Definition for customs func.
def format_values(value):
    """Formating values in a way that will replace entries of "," with ".". Also let you repair input if it was used as wrong type

    Args:
        value (not_specified): User input which than is evaulated and ideally transfered to float

    Returns:
        value: Return formated value of the entry input
    """
    try:
        value = float(value)
        return value
    except ValueError:
        try:
            value = value.replace(",",".")
            value = float(value)
            return value
        except ValueError:
            while True:
                try:
                    value = input("Invalid input, Please enter a value, not a text: ")
                    value = value.replace(",",".")
                    value = float(value)
                    return value
                except:
                    continue

def Repair_table(used_df, repair_it):
    """Repair function which will specify the dimensions of the matrix and then rewrite it´s value

    Args:
        used_df (pandas.core.frame.DataFrame): used dataframe to repair within
        repair_it (str): string of responses 

    Returns:
        _type_: _description_
    """
    if repair_it.lower() in response:
        print("\n\n\n------------Actual REPAIRED table ----------")
        print(used_df)
        print("\n Now tell me dimensions which you wish to change")
        first_dim = int(input("Tell me first dimension of the matrix - i.e. order num. of the row: "))
        second_dim = int(input("Tell me second dimension of the matrix - i.e. order num. of the column: "))
        value = format_values(input("What value you would like to write there: "))
        used_df.iat[first_dim,second_dim] = value
        print("\n")
        print(used_df)
        return True
    else:
        return False
    
    
start_pack = input("Do you want to add packing manually ? y/n: ")
if start_pack:
            
    weight_inp = input("Do you wish to include weight columns? y/n: ")        
    while True:
        # Get user input for dimensions
        quant = format_values(input("Enter how many cartons: "))
        length = format_values(input("Enter length: "))
        width = format_values(input("Enter width: "))
        height = format_values(input("Enter height: "))
                
        if weight_inp in response:
            weight = format_values(input("Enter weight: "))
            

        # Create new row as dictionary and then convert it to dataframe which can be concatenate afterwards with existing dataframe
        if weight_inp not in response:
            new_row = {"Quantity [0]":quant,"Length [1]":length,"Width [2]":width,"Height [3]": height}
            new_row = pd.DataFrame([new_row])
            df = pd.concat([df, new_row], ignore_index=True)
            print(df)
        if weight_inp in response:
            new_row_weight = {"Quantity [0]":quant,"Length [1]":length,"Width [2]":width,"Height [3]": height,"Weight [4]": weight, "----": []}
            new_row_weight = pd.DataFrame([new_row_weight])
            df2 = pd.concat([df2, new_row_weight], ignore_index=True)
            print(df2)


        more = input("Do you want to add more dimensions? y/n: ")
        if  more.lower() not in response:
            break


        
        
    repair_it = input("Do you wish to repair any value? y/n: ")
    flag = False
    if weight_inp in response:
        flag = Repair_table(df2,repair_it)
    elif weight_inp not in response:
        flag = Repair_table(df,repair_it)
    
    while flag:
        repair_it = input("Do you wish to repair any value? y/n: ")
        if weight_inp in response:
            flag = Repair_table(df2,repair_it)
        elif weight_inp not in response:
            flag = Repair_table(df,repair_it)
        
    def add_packing(row):
        """Function which will evaulate wheter is the dimensions provided suitable more for pallet or carton

        Args:
            row (row): each row of the df

        Returns:
            str: what kind of string to apply to Packing column as output
        """
        if (row["Length [1]"] >= 80 and row["Width [2]"] >= 60) or (row["Length [1]"] >= 60 and row["Width [2]"] >= 80):
            return "plt"
        else:
            return "ctn"
        
    if weight_inp not in response:
        df_orig = pd.DataFrame.copy(df)
        df_orig["Packing"] = df_orig.apply(add_packing, axis=1)
        df["Volume"] = df["Quantity [0]"] * df["Length [1]"] * df["Width [2]"] * df["Height [3]"] / 1000000
        df["Volumetric weight (167*cbm)"] = df["Length [1]"] * df["Width [2]"] * df["Height [3]"] * 167 / 1000000
        columns_for_sum =["Quantity [0]","Volume","Volumetric weight (167*cbm)"]

        total = df.loc["TOTAL"] = df[columns_for_sum].sum(numeric_only= True, axis = 0, skipna = True)
        
        
        df_orig = df_orig.applymap(lambda x: int(x) if type(x) == float and x == round(x) else x)
        #new column and conversion to one line which is printable
        df_orig["len_Wi_Hei_Pack"] = df_orig.apply(lambda x: f'{x["Quantity [0]"]} {x["Packing"]}  {x["Length [1]"]}x{x["Width [2]"]}x{x["Height [3]"]} cm ', axis=1)
        
        
        

    
    
        
    if weight_inp in response:
        df2_orig = pd.DataFrame.copy(df2)
        df2["Packing"] = df2.apply(add_packing, axis=1)
        df2["Volume"] = df2["Quantity [0]"] * df2["Length [1]"] * df2["Width [2]"] * df2["Height [3]"] / 1000000
        df2["Total Weight"] = df2["Quantity [0]"]*df2["Weight [4]"]
        df2["Volumetric weight (167*cbm)"] = df2["Volume"] * 167

        columns_for_sum =["Quantity [0]","Volume","Total Weight","Volumetric weight (167*cbm)"]
        total = df2.loc["TOTAL"] = df2[columns_for_sum].sum(numeric_only= True, axis = 0, skipna = True)

   
        df2_orig = df2_orig.applymap(lambda x: int(x) if type(x) == float and x == round(x) else x)
        #new column and conversion to one line which is printable
        df2_orig["len_Wi_Hei_Wei"] = df2_orig.apply(lambda x: f'{x["Quantity [0]"]}x  {x["Length [1]"]}x{x["Width [2]"]}x{x["Height [3]"]} cm  {x["Weight [4]"]} kg/', axis=1) 
        df2["Packing"] = df2.apply(lambda x: f'{x["Packing"]}',axis=1 )
        df2_orig = df2_orig.assign(Packing=df2["Packing"], blank1=["----"]* len(df2_orig), blank2=["----"]* len(df2_orig), Volume=df2["Volume"], Weight=df2["Weight [4]"])
        df2_orig["len_Wi_Hei_Wei_Pack"] = df2_orig.apply(lambda x: f'{x["Quantity [0]"]}x  {x["Length [1]"]}x{x["Width [2]"]}x{x["Height [3]"]} cm  {x["Weight [4]"]} kg/{x["Packing"]}', axis=1)
        df2_orig["Total Weight"] = df2_orig["Quantity [0]"] * df2_orig["Weight [4]"]
        cols_for_sum =["Volume","Total Weight"]
        total = df2_orig.loc["TOTAL"] = df2_orig[cols_for_sum].sum(numeric_only= True, axis = 0, skipna = True)
        df2_orig[["Volume","Total Weight","----","----","len_Wi_Hei_Wei_Pack"]].to_excel("vol2.xlsx", index=False, sheet_name='Sheet1', header=True)
   



#-----------------------------------------------------------------------------
#-----------------------------------------------------------------------------
#-----------------------------------------------------------------------------

#EXCEL PART

workbook = load_workbook("PL_blank.xlsx")
ws = workbook["List1"]



adress_cell = ws["B9"]
adress_cell.alignment = Alignment(horizontal='left',vertical="top", wrap_text=True)
adress_cell.value = delivery_adress


order_cell = ws["A3"]
order_cell.value = "ORDER : " +  ", ".join(order_num)

PL_cell = ws["A5"]
#PL_cell.alignment = Alignment(horizontal='center', wrap_text=True)
PL_cell.value = "PACKING LIST - " + ", ".join(dl_note)


INV_cell = ws["A7"]
#INV_cell.alignment = Alignment(horizontal='center', wrap_text=True)
INV_cell.value = "INVOICE: " + invoice

inco_cell = ws["A36"]
inco_cell.value = founded_inco + ": " + founded_city + " (according INCOTERMS 2010)"


def write_description(worksheet, column, values, start_row):
    for i, value in enumerate(values):
        row = start_row + i
        cell = worksheet[column + str(row)]
        cell.value = value
        if cell.value.isdigit():
            cell.value = int(cell.value)


#packing
def count_packing(df):
    plt_count = df.loc[df['Packing'] == 'plt', 'Quantity [0]'].sum()
    ctn_count = df.loc[df['Packing'] == 'ctn', 'Quantity [0]'].sum()
    if plt_count == 0:
        if ctn_count == 1:
           return f'{int(ctn_count)} ctn' 
        else:
            return f'{int(ctn_count)} ctns' 
    elif ctn_count == 0:
        if plt_count == 1:
           return f'{int(plt_count)} plt' 
        else:
            return f'{int(plt_count)} plts' 
    elif plt_count > 1 and ctn_count == 1:
        return f'{int(plt_count)} plts and {int(ctn_count)} ctn'
    elif plt_count == 1 and ctn_count > 1:
        return f'{int(plt_count)} plt and {int(ctn_count)} ctns'
    else:
        return f'{int(plt_count)} plts and {int(ctn_count)} ctns'
    
pack_cell = ws["E20"]
weight_cell = ws["C46"]
unit_cell = ws["D46"]
cbm_desc = ws["B47"]
cbm_tot = ws["C47"]
cbm_unit = ws["D47"]
quant_and_kind = ws["C45"]
header_tot_w = ws["E41"]

if start_pack and weight_inp in response:
    for i, value in enumerate(df2_orig["len_Wi_Hei_Wei_Pack"]):
        cell = ws.cell(row=pack_cell.row + i, column=pack_cell.column)
        cell.value = value

    weight_cell.value = df2_orig.loc["TOTAL", "Total Weight"] 
    weight_cell.alignment = Alignment(horizontal='right', wrap_text = True)
    unit_cell.value = "Kgs"
    unit_cell.alignment = Alignment(horizontal="left", wrap_text = True)
    cbm_desc.value = "Total cbm:"
    cbm_desc.alignment = Alignment(horizontal="left", wrap_text = True)
    cbm_tot.value = df2_orig.loc["TOTAL", "Volume"] 
    cbm_unit.value = "cbm"
    packing = count_packing(df2_orig)
    quant_and_kind.value = packing
    header_w = round(df2_orig.loc["TOTAL", "Total Weight"], 2)
    header_tot_w.value = "Total Gross Weight: " + "{:.2f}".format(header_w).replace(".", ",") + " Kgs"
    
elif start_pack and weight_inp not in response:
    for i, value in enumerate(df_orig["len_Wi_Hei_Pack"]):
        cell = ws.cell(row=pack_cell.row + i, column=pack_cell.column)
        cell.value = value
    packing = count_packing(df_orig)
    quant_and_kind.value = packing
    cbm_desc.value = "Total cbm:"
    cbm_desc.alignment = Alignment(horizontal="left", wrap_text = True)
    cbm_tot.value = df.loc["TOTAL", "Volume"] 
    cbm_unit.value = "cbm"
    
write_description(ws,"A",codes,20)
write_description(ws,"B",description_list,20)
write_description(ws,"C",quantity_list,20)
pcs =[]
count = len(codes)
for i in range(count):
    pcs.append("pcs")
write_description(ws,"D",pcs,20)

workbook.save(f"PL {dl_note[0]} {founded_city} {founded_state}.xlsx")
print("PL was saved and exported")



