import re
from pdfminer.high_level import extract_pages, extract_text
import geonamescache
import pycountry
import pypdf
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment

gc = geonamescache.GeonamesCache()
cities = gc.get_cities()
state_list = list(pycountry.countries) 
city_names = set(city['name'] for city in cities.values())

"""for page_layout in extract_pages("invoice.pdf"):
    for element in page_layout:
        print(element)"""



while True:
    try:
        ask_for_name = input("What pdf file do you want to load: ")
        inv_header = ask_for_name + ".pdf"        
        text = extract_text(inv_header)
        break
    except Exception as chyba:
        print(chyba)


delivery = text.find("Delivery")
def end_of_delivery():
    end1 = text.find("Payment")
    end2 = text.find("For\xa0custom")
    len1 = len(text[delivery:end1-1])
    len2 = len(text[delivery:end2-1])
    if len1 < len2 and len1 > 50:
        print("Returning end1 - Payment")
        return end1
    elif len2 < len1 and len2 > 50:
        print("Returning end2 - For custom")
        return end2
    else:
        return end1
end = end_of_delivery()    


delivery_adress = text[delivery:end-1]
delivery_adress = delivery_adress.replace("Delivery adress:","")
delivery_adress = delivery_adress.lstrip("\n")

incoterm_list = ["EXW","FCA","FOB","CFR","CIF","CIP""CPT","DAP","DAT","DDP"]

def incoterm_overlook(incoterm_list):
    for i in incoterm_list:
        found = int(text.find(i))
        if found != -1:
            if i == "DDP" and text[found+4] == "V":
                return "DDP VAT Unpaid"
            else:
                return i

def find_city(delivery_adress):
    tokenized = delivery_adress.split()
    if "Asahi" in tokenized:
        tokenized.remove("Asahi")
    for i in tokenized:
        if gc.get_cities_by_name(i):
            return i
    
    

def find_state_code():
    tokenized = delivery_adress.split()
    if "AS" in tokenized:
        tokenized.remove("AS")
    for i in tokenized:
        if i in [country.alpha_2 for country in pycountry.countries]:
            return i 
      
    
founded_inco = incoterm_overlook(incoterm_list)
founded_city = find_city(delivery_adress)
founded_state = find_state_code()


pdf_reader = pypdf.PdfReader(inv_header)



# Get the number of pages in the PDF file
num_pages = len(pdf_reader.pages)

# Loop over each page in the PDF file
def print_pages():
    for page_num in range(num_pages):
        # Get the page object for the current page
        page = pdf_reader.pages[page_num]

        # Extract the text content from the page
        page_text = page.extract_text()

        # Print the text content to the console
        if page_text.find("PRODUCTS") == -1:
            table_start = page_text.find("GOODS")
        else:
            table_start = page_text.find("PRODUCTS")
        table_end = page_text.find("OTHERS")
        table = page_text[table_start:table_end]
        return table,page_text
    



# Define the regular expression patterns
zb_pattern = r"ZB\w+"
quant_pattern = r"pc\s*\d{1,3}(?:\s+\d{3})*" # matches zero or more groups of one or more whitespace characters followed by exactly three digits.
desc_pattern = r"%\s*\d{1,3}(?:\s+\d{3})*,\d{2}" # This pattern matches a percentage sign followed by a space, and then 1 to 5 digits, a comma, and 2 more digits. It can be used to match strings like "% 465,12".
netto_pattern = r"%\s*\d{1,3}(?:\s+\d{3})*,\d{2}" # + ,00
brutto_pattern = r"(?<=Packaging:).*?\n(\d[\d,. ]*)\s*Total Gross Weight"
inv_pattern = r"\d{2}03\w+"  
dl_pattern = r"\d{2}(?:SL|MH)01\w+"
order_pattern = r"\d{2}[A-Z0-9]{4}0\d{9}INVOICE"


table,page_text = print_pages()

print(table)
# Find all matches of the code pattern in the input string
codes = re.findall(zb_pattern, table)
quantity_list = re.findall(quant_pattern, table)
quantity_list = [quantity.replace('pc', '').replace('\xa0', '').strip() for quantity in quantity_list]
brutto_list = re.findall(brutto_pattern,page_text)
description = re.findall(desc_pattern, table)
#description = [desc.replace("\xa0"," ").strip() for desc in description] # this will remove \xa0 from the string but it will not find it afterwards
netto = re.findall(netto_pattern,table)
netto = [net.replace("%","").replace('\xa0', '') for net in netto]


def inv_lookup(what_to_look_for, text):
    findings = re.findall(what_to_look_for, text)
    return findings

invoice = re.search(inv_pattern, page_text)
invoice = invoice.group(0)

dl_note = inv_lookup(dl_pattern,page_text)
order_num = inv_lookup(order_pattern, page_text)
order_num = [order.replace("INVOICE","") for order in order_num]

def last_element(table, pattern):
    row_table = table.split('\n')
    end_list = []
    for row in row_table:
        if re.search(pattern, row):
            end = row.split()[-1]
            end_list.append(end)
    return end_list, row_table

end_list,row_table = last_element(table,zb_pattern)

print("\n\n\n")
def solidity_check(founded_city, founded_state):
    if founded_city is None or founded_state is None:
        if founded_city is None:
            print("City is missing")
            founded_city = input("Type the name of the city: ")
        if founded_state is None:
            print("Country code is missing")
            founded_state = input("Type the code of the country: ")
        return solidity_check(founded_city, founded_state)
    
    print("City and Country code are OK")
    return founded_city, founded_state
founded_city,founded_state = solidity_check(founded_city,founded_state)


# Print the results
print("\n\n\n")
print("Founded incoterm: ",founded_inco)
print("Founded city: ",founded_city)
print("Founded state: ",founded_state)
print("Codes: ", codes)
print("Quantity: ", quantity_list)
print("Start Product description: ", description)
print("Netto list is: ", netto)
print("Brutto list is: ", brutto_list)
print("End list : ",end_list)  
print("Invoice number is: ", invoice)
print("Delivery note number is:", dl_note)
print("Order number is: ", order_num)

#perc_netto = ["%" + x for x in netto] # can replicate like ["%" + x,   for x in netto]
#print("percent netto: ",perc_netto)

def find_description():
    description_list = []
    for start, end in zip(description, end_list):
        start_pos = table.find(start) + len(start) + 1
        end_pos = table.find(end, start_pos) # finds the index of the first occurrence of end after the start substring.
        desc_row=table[start_pos:end_pos + 2]
        description_list.append(desc_row)
    return description_list

description_list = find_description()
print(description_list)
        
print("\n\n\n" )
#print(page_text)
print(text)
print(repr(delivery_adress))












#EXCEL PART

workbook = load_workbook("PL_blank.xlsx")
ws = workbook["List1"]



adress_cell = ws["B9"]
adress_cell.alignment = Alignment(horizontal='left',vertical="top", wrap_text=True)
adress_cell.value = delivery_adress


order_cell = ws["A3"]
order_cell.value = "ORDER : " +  ", ".join(order_num)

PL_cell = ws["A5"]
#PL_cell.alignment = Alignment(horizontal='center', wrap_text=True)
PL_cell.value = "PACKING LIST - " + ", ".join(dl_note)


INV_cell = ws["A7"]
#INV_cell.alignment = Alignment(horizontal='center', wrap_text=True)
INV_cell.value = "INVOICE: " + invoice

inco_cell = ws["A36"]
inco_cell.value = founded_inco + ": " + founded_city + " (according INCOTERMS 2010)"


def write_description(worksheet, column, values, start_row):
    for i, value in enumerate(values):
        row = start_row + i
        cell = worksheet[column + str(row)]
        cell.value = value
        if cell.value.isdigit():
            cell.value = int(cell.value)

write_description(ws,"A",codes,20)
write_description(ws,"B",description_list,20)
write_description(ws,"C",quantity_list,20)
pcs =[]
count = len(codes)
for i in range(count):
    pcs.append("pcs")
write_description(ws,"D",pcs,20)

workbook.save(f"PL {dl_note[0]} {founded_city} {founded_state}.xlsx")
print("PL was saved and exported")



